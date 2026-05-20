#!/usr/bin/env python3
"""
AWS Access Review — PULL (read-only)
=====================================
Consumes AWS-NATIVE governance signals and produces a reviewer-ready xlsx
workbook with one row per principal/finding and a locked decision dropdown.

This script READS ONLY. It never modifies IAM. Remediation is handled by the
separate, explicitly-gated apply script after a human completes the review.

Native sources consumed
------------------------
  * IAM Access Analyzer — Unused Access findings (org-level analyzer)
        unused roles, unused users, unused permissions (90-day window)
  * IAM Credential Reports (per account)
        user inventory, key age, last-used, MFA status
  * IAM Access Advisor / last-accessed (per principal, optional deep mode)

Scope
-----
Org-level. Iterates member accounts via AWS Organizations and assumes a
read-only review role in each. Falls back to the current account if --single.

Usage
-----
  python aws_access_review_pull.py \
      --analyzer-arn arn:aws:access-analyzer:us-east-1:111111111111:analyzer/org-unused-access \
      --review-role OrganizationAccountAccessRole-ReadOnly \
      --out aws_access_review.xlsx

  python aws_access_review_pull.py --single --out aws_access_review.xlsx

Requirements
------------
  pip install boto3 openpyxl
  Read-only permissions: access-analyzer:ListFindings, iam:GenerateCredentialReport,
  iam:GetCredentialReport, organizations:ListAccounts, sts:AssumeRole,
  iam:GetServiceLastAccessedDetails (deep mode only).
"""

import argparse
import base64
import csv
import datetime as dt
import io
import sys
from collections import defaultdict

try:
    import boto3
    from botocore.exceptions import ClientError, BotoCoreError
except ImportError:
    sys.exit("boto3 is required:  pip install boto3")

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")


DECISION_OPTIONS = ["", "APPROVE", "REVOKE", "INVESTIGATE"]
UNUSED_AGE_LABEL = "90d"


# --------------------------------------------------------------------------- #
# Native source collectors
# --------------------------------------------------------------------------- #
def collect_unused_access_findings(session, analyzer_arn):
    """Pull IAM Access Analyzer unused-access findings (the primary signal)."""
    client = session.client("accessanalyzer")
    findings = []
    paginator = client.get_paginator("list_findings_v2")
    try:
        for page in paginator.paginate(analyzerArn=analyzer_arn):
            for f in page.get("findings", []):
                findings.append({
                    "finding_type": f.get("findingType", "Unknown"),
                    "resource": f.get("resource", ""),
                    "resource_owner": f.get("resourceOwnerAccount", ""),
                    "status": f.get("status", ""),
                    "analyzed_at": str(f.get("analyzedAt", "")),
                })
    except (ClientError, BotoCoreError) as e:
        print(f"  ! Access Analyzer query failed: {e}", file=sys.stderr)
    return findings


def collect_credential_report(session, account_id):
    """Generate + fetch the IAM credential report for one account."""
    client = session.client("iam")
    try:
        client.generate_credential_report()
    except ClientError:
        pass  # report may already be generating / fresh
    # Poll briefly for the report
    import time
    report_csv = None
    for _ in range(10):
        try:
            resp = client.get_credential_report()
            report_csv = resp["Content"].decode("utf-8")
            break
        except ClientError as e:
            if "ReportNotPresent" in str(e) or "ReportInProgress" in str(e):
                time.sleep(2)
                continue
            print(f"  ! credential report failed ({account_id}): {e}", file=sys.stderr)
            return []
    if not report_csv:
        return []

    rows = []
    reader = csv.DictReader(io.StringIO(report_csv))
    for r in reader:
        rows.append({
            "account": account_id,
            "user": r.get("user", ""),
            "arn": r.get("arn", ""),
            "mfa_active": r.get("mfa_active", ""),
            "password_last_used": r.get("password_last_used", ""),
            "access_key_1_active": r.get("access_key_1_active", ""),
            "access_key_1_last_used_date": r.get("access_key_1_last_used_date", ""),
            "access_key_2_active": r.get("access_key_2_active", ""),
            "access_key_2_last_used_date": r.get("access_key_2_last_used_date", ""),
        })
    return rows


def list_org_accounts(session):
    """Enumerate active member accounts via AWS Organizations."""
    client = session.client("organizations")
    accounts = []
    try:
        paginator = client.get_paginator("list_accounts")
        for page in paginator.paginate():
            for a in page["Accounts"]:
                if a["Status"] == "ACTIVE":
                    accounts.append({"id": a["Id"], "name": a["Name"]})
    except (ClientError, BotoCoreError) as e:
        print(f"  ! Organizations list failed (running single-account?): {e}",
              file=sys.stderr)
    return accounts


def assume_role(base_session, account_id, role_name):
    """Assume a read-only review role in a member account."""
    sts = base_session.client("sts")
    role_arn = f"arn:aws:iam::{account_id}:role/{role_name}"
    try:
        creds = sts.assume_role(
            RoleArn=role_arn,
            RoleSessionName="AccessReviewPull",
        )["Credentials"]
        return boto3.Session(
            aws_access_key_id=creds["AccessKeyId"],
            aws_secret_access_key=creds["SecretAccessKey"],
            aws_session_token=creds["SessionToken"],
        )
    except (ClientError, BotoCoreError) as e:
        print(f"  ! assume-role failed for {account_id}: {e}", file=sys.stderr)
        return None


# --------------------------------------------------------------------------- #
# Workbook builder
# --------------------------------------------------------------------------- #
def build_workbook(findings, cred_rows, out_path):
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    decision_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    def add_decision_validation(ws, col_letter, nrows):
        dv = DataValidation(
            type="list",
            formula1='"APPROVE,REVOKE,INVESTIGATE"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.error = "Choose APPROVE, REVOKE, or INVESTIGATE."
        dv.errorTitle = "Invalid decision"
        dv.prompt = "Select a review decision."
        dv.promptTitle = "Reviewer decision"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{nrows + 1}")

    # ---- Sheet 1: Unused Access (Access Analyzer) ----
    ws1 = wb.active
    ws1.title = "Unused Access"
    cols1 = ["Account", "Finding Type", "Resource (Principal)", "Native Signal",
             "Status", "Analyzed", "Reviewer Decision", "Reviewer", "Date", "Justification"]
    ws1.append(cols1)
    style_header(ws1, len(cols1))
    for f in findings:
        ws1.append([
            f["resource_owner"], f["finding_type"], f["resource"],
            f"Unused {UNUSED_AGE_LABEL}", f["status"], f["analyzed_at"],
            "", "", "", "",
        ])
    if findings:
        add_decision_validation(ws1, "G", len(findings))
        for r in range(2, len(findings) + 2):
            ws1.cell(row=r, column=7).fill = decision_fill
    widths1 = [16, 22, 46, 14, 10, 22, 18, 16, 12, 36]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 2: Credential Inventory ----
    ws2 = wb.create_sheet("Credential Inventory")
    cols2 = ["Account", "User", "ARN", "MFA", "Pwd Last Used",
             "Key1 Active", "Key1 Last Used", "Key2 Active", "Key2 Last Used",
             "Reviewer Decision", "Reviewer", "Date", "Justification"]
    ws2.append(cols2)
    style_header(ws2, len(cols2))
    for r in cred_rows:
        ws2.append([
            r["account"], r["user"], r["arn"], r["mfa_active"],
            r["password_last_used"], r["access_key_1_active"],
            r["access_key_1_last_used_date"], r["access_key_2_active"],
            r["access_key_2_last_used_date"], "", "", "",
        ])
    if cred_rows:
        add_decision_validation(ws2, "J", len(cred_rows))
        for r in range(2, len(cred_rows) + 2):
            ws2.cell(row=r, column=10).fill = decision_fill
    widths2 = [14, 18, 40, 8, 18, 11, 18, 11, 18, 18, 14, 12, 30]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 3: Summary / instructions ----
    ws3 = wb.create_sheet("README", 0)
    ws3.column_dimensions["A"].width = 100
    lines = [
        ("AWS Access Review — Reviewer Workbook", True),
        ("", False),
        (f"Generated: {dt.datetime.now(dt.timezone.utc).isoformat()} (read-only pull)", False),
        ("", False),
        ("How to complete this review:", True),
        ("1. Review each row on the 'Unused Access' and 'Credential Inventory' tabs.", False),
        ("2. Set 'Reviewer Decision' to APPROVE, REVOKE, or INVESTIGATE.", False),
        ("3. Fill in your name (Reviewer), the Date, and a short Justification.", False),
        ("4. Every row must have a decision — no blanks — before the review is Done.", False),
        ("5. Save and return this file. REVOKE rows are actioned by the gated apply step.", False),
        ("", False),
        ("Native signals used:", True),
        ("  • Unused Access tab  → IAM Access Analyzer unused-access findings (90-day).", False),
        ("  • Credential tab     → IAM credential report (key age, last-used, MFA).", False),
        ("", False),
        ("This workbook does not change any access. It is evidence + a decision record.", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        c = ws3.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=13 if i == 1 else 11,
                          color="1F3864" if i == 1 else "000000")

    wb.save(out_path)


# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="AWS access review pull (read-only).")
    ap.add_argument("--analyzer-arn", help="Org unused-access analyzer ARN.")
    ap.add_argument("--review-role", default="OrganizationAccountAccessRole-ReadOnly",
                    help="Read-only role to assume in member accounts.")
    ap.add_argument("--single", action="store_true",
                    help="Single-account mode (skip Organizations enumeration).")
    ap.add_argument("--out", default="aws_access_review.xlsx")
    ap.add_argument("--profile", default=None, help="AWS profile name.")
    args = ap.parse_args()

    base = boto3.Session(profile_name=args.profile) if args.profile else boto3.Session()

    print("AWS Access Review — PULL (read-only)")
    findings = []
    if args.analyzer_arn:
        print("· Pulling Access Analyzer unused-access findings ...")
        findings = collect_unused_access_findings(base, args.analyzer_arn)
        print(f"  {len(findings)} unused-access findings")
    else:
        print("· No --analyzer-arn given; skipping Access Analyzer (recommended to enable).")

    cred_rows = []
    if args.single:
        acct = base.client("sts").get_caller_identity()["Account"]
        print(f"· Single account {acct}: credential report ...")
        cred_rows = collect_credential_report(base, acct)
    else:
        accounts = list_org_accounts(base)
        print(f"· {len(accounts)} active org accounts")
        for a in accounts:
            sess = assume_role(base, a["id"], args.review_role)
            if not sess:
                continue
            print(f"  - {a['name']} ({a['id']}): credential report ...")
            cred_rows.extend(collect_credential_report(sess, a["id"]))

    print(f"· Building workbook: {args.out}")
    build_workbook(findings, cred_rows, args.out)
    print(f"Done. {len(findings)} findings, {len(cred_rows)} credential rows.")
    print("Reviewer fills decisions; REVOKE rows feed the gated apply step.")


if __name__ == "__main__":
    main()
