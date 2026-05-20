#!/usr/bin/env python3
"""
Access Review — APPLY (gated remediation)
==========================================
Reads a COMPLETED reviewer workbook (AWS or GCP) and remediates ONLY the rows a
human explicitly marked "REVOKE". Everything about this step is deliberately
conservative:

  * DRY-RUN BY DEFAULT. It prints exactly what it would do and changes nothing.
  * Acts on REVOKE rows ONLY. APPROVE / INVESTIGATE / blank are never touched.
  * Refuses to run unless every in-scope row has a decision (no blanks).
  * Requires an explicit --apply flag AND an interactive typed confirmation
    before a single IAM call is made.
  * Writes an immutable evidence log (who/what/when/result) for the auditor.

>>> v1 STATUS: the per-cloud revoke calls are intentionally STUBBED. <<<
The plumbing — parsing, gating, validation, evidence logging — is complete and
real. The actual destructive IAM calls are marked TODO so that this can be
reviewed and approved by the customer before it is ever allowed to write to IAM.
This matches the "show the review sheet first, trust it, then enable apply"
posture that customers expect.

Usage
-----
  # Always start here — shows the plan, changes nothing:
  python access_review_apply.py --cloud aws --workbook aws_access_review.xlsx

  # Only after review sign-off, and only when stubs are implemented + tested:
  python access_review_apply.py --cloud aws --workbook aws_access_review.xlsx --apply

Requirements
------------
  pip install openpyxl boto3 google-cloud-iam   (clouds as needed)
"""

import argparse
import datetime as dt
import json
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")


DECISION_COL_HEADER = "Reviewer Decision"
VALID_DECISIONS = {"APPROVE", "REVOKE", "INVESTIGATE"}


# --------------------------------------------------------------------------- #
# Workbook parsing + validation
# --------------------------------------------------------------------------- #
def _find_decision_col(ws):
    for c in range(1, ws.max_column + 1):
        if (ws.cell(row=1, column=c).value or "").strip() == DECISION_COL_HEADER:
            return c
    return None


def load_decisions(workbook_path):
    """Return (rows, blanks) where rows is a list of dicts with decision + context."""
    wb = load_workbook(workbook_path, data_only=True)
    rows = []
    blanks = 0
    for ws in wb.worksheets:
        if ws.title in ("README",):
            continue
        dcol = _find_decision_col(ws)
        if not dcol:
            continue  # not a review sheet (e.g. Recommender Insights detail tab)
        headers = [(ws.cell(row=1, column=c).value or f"col{c}")
                   for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            decision = (ws.cell(row=r, column=dcol).value or "").strip().upper()
            # Skip fully empty rows
            if not any(ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)):
                continue
            if decision == "":
                blanks += 1
                continue
            if decision not in VALID_DECISIONS:
                blanks += 1
                continue
            record = {headers[c - 1]: ws.cell(row=r, column=c).value
                      for c in range(1, ws.max_column + 1)}
            record["_sheet"] = ws.title
            record["_row"] = r
            record["_decision"] = decision
            rows.append(record)
    return rows, blanks


# --------------------------------------------------------------------------- #
# Remediation stubs (intentionally non-destructive in v1)
# --------------------------------------------------------------------------- #
def revoke_aws(record, dry_run):
    """
    TODO(v2): implement the actual revoke. Depending on the sheet:
      - Unused Access (role/user): detach policy / delete unused permission, or
        deactivate access keys via iam:UpdateAccessKey(Status='Inactive').
      - Credential Inventory: deactivate stale access keys / disable login profile.
    Must be least-destructive first (deactivate before delete) and reversible.
    """
    target = record.get("Resource (Principal)") or record.get("ARN") or record.get("User")
    action = f"[AWS] would revoke/deactivate: {target}"
    if dry_run:
        return ("DRY-RUN", action)
    # --- real call goes here, gated behind --apply and confirmation ---
    return ("SKIPPED-STUB", action + "  (stub not yet implemented)")


def revoke_gcp(record, dry_run):
    """
    TODO(v2): implement removeIamPolicyBinding equivalent:
      resource = record['Resource']; role = record['Role']; member = record['Member']
      Use the Resource Manager / per-service setIamPolicy with read-modify-write,
      removing exactly {member, role}. Always re-read the live policy first to
      avoid clobbering concurrent changes (etag check).
    """
    target = f"{record.get('Member')} :: {record.get('Role')} on {record.get('Resource')}"
    action = f"[GCP] would remove binding: {target}"
    if dry_run:
        return ("DRY-RUN", action)
    # --- real call goes here, gated behind --apply and confirmation ---
    return ("SKIPPED-STUB", action + "  (stub not yet implemented)")


REVOKERS = {"aws": revoke_aws, "gcp": revoke_gcp}


# --------------------------------------------------------------------------- #
def write_evidence_log(cloud, results, blanks, path):
    log = {
        "generated_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
        "cloud": cloud,
        "rows_with_blank_or_invalid_decision": blanks,
        "actions": [{"status": s, "detail": d} for s, d in results],
        "summary": {
            "total_actions": len(results),
            "dry_run": all(s == "DRY-RUN" for s, _ in results) if results else True,
        },
    }
    with open(path, "w") as f:
        json.dump(log, f, indent=2)
    return path


def main():
    ap = argparse.ArgumentParser(description="Gated access-review remediation.")
    ap.add_argument("--cloud", required=True, choices=["aws", "gcp"])
    ap.add_argument("--workbook", required=True, help="Completed reviewer workbook (.xlsx).")
    ap.add_argument("--apply", action="store_true",
                    help="Actually apply (default is dry-run). Still requires typed confirm.")
    ap.add_argument("--evidence-out", default=None,
                    help="Path for the JSON evidence log.")
    ap.add_argument("--allow-blanks", action="store_true",
                    help="Override the no-blank-decisions safety check (NOT recommended).")
    args = ap.parse_args()

    print(f"Access Review — APPLY ({args.cloud.upper()})  [{'APPLY' if args.apply else 'DRY-RUN'}]")
    rows, blanks = load_decisions(args.workbook)

    revoke_rows = [r for r in rows if r["_decision"] == "REVOKE"]
    approve = sum(1 for r in rows if r["_decision"] == "APPROVE")
    investigate = sum(1 for r in rows if r["_decision"] == "INVESTIGATE")

    print(f"· Decisions: {approve} APPROVE, {len(revoke_rows)} REVOKE, "
          f"{investigate} INVESTIGATE, {blanks} blank/invalid")

    # Safety gate 1: no incomplete reviews
    if blanks and not args.allow_blanks:
        sys.exit(f"REFUSING: {blanks} row(s) have no valid decision. "
                 f"A review must be complete before remediation. "
                 f"Use --allow-blanks only if you understand the risk.")

    if not revoke_rows:
        print("Nothing marked REVOKE. Nothing to do.")
        write_evidence_log(args.cloud, [], blanks,
                           args.evidence_out or f"{args.cloud}_apply_evidence.json")
        return

    dry_run = not args.apply
    revoker = REVOKERS[args.cloud]

    # Safety gate 2: explicit typed confirmation before any apply
    if not dry_run:
        print(f"\nAbout to REVOKE {len(revoke_rows)} item(s) in {args.cloud.upper()}.")
        confirm = input('Type "REVOKE" exactly to proceed: ').strip()
        if confirm != "REVOKE":
            sys.exit("Confirmation failed. No changes made.")

    results = []
    for r in revoke_rows:
        status, detail = revoker(r, dry_run)
        results.append((status, detail))
        print(f"  {status}: {detail}")

    out = args.evidence_out or f"{args.cloud}_apply_evidence.json"
    write_evidence_log(args.cloud, results, blanks, out)
    print(f"\nEvidence log written: {out}")
    if dry_run:
        print("DRY-RUN complete. No access was changed. Re-run with --apply to act "
              "(once the v2 revoke stubs are implemented and tested).")


if __name__ == "__main__":
    main()
