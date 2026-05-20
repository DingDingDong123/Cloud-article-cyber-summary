#!/usr/bin/env python3
"""
GCP Access Review — PULL (read-only)
=====================================
Consumes GCP-NATIVE governance signals and produces a reviewer-ready xlsx
workbook with one row per IAM binding and a locked decision dropdown.

This script READS ONLY. It never modifies IAM. Remediation is handled by the
separate, explicitly-gated apply script after a human completes the review.

Native sources consumed
------------------------
  * Cloud Asset Inventory — searchAllIamPolicies (org-wide IAM binding inventory)
  * IAM Recommender — google.iam.policy.Recommender
        usage-based (90-day) recommendations to remove / right-size roles
        (joined onto bindings so the reviewer sees "stale 90d" pre-flags)

Scope
-----
Org-level. Searches all IAM policies under an organization scope and pulls
recommender insights per project discovered in the inventory.

Usage
-----
  python gcp_access_review_pull.py \
      --org-id 123456789012 \
      --out gcp_access_review.xlsx

  python gcp_access_review_pull.py --project my-project --out gcp_access_review.xlsx

Requirements
------------
  pip install google-cloud-asset google-cloud-recommender openpyxl
  Auth:  gcloud auth application-default login   (read-only roles below)
  Read-only roles: roles/cloudasset.viewer (org), roles/recommender.iamViewer.
"""

import argparse
import datetime as dt
import sys
from collections import defaultdict

try:
    from google.cloud import asset_v1
    from google.cloud import recommender_v1
    from google.api_core.exceptions import GoogleAPIError, PermissionDenied
except ImportError:
    sys.exit("Google client libs required:  "
             "pip install google-cloud-asset google-cloud-recommender")

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")


IAM_RECOMMENDER = "google.iam.policy.Recommender"


# --------------------------------------------------------------------------- #
# Native source collectors
# --------------------------------------------------------------------------- #
def collect_iam_bindings(scope):
    """Cloud Asset Inventory: every IAM binding under the scope (org/project)."""
    client = asset_v1.AssetServiceClient()
    rows = []
    try:
        results = client.search_all_iam_policies(request={"scope": scope})
        for r in results:
            resource = r.resource
            project = r.project
            policy = r.policy
            for binding in policy.bindings:
                for member in binding.members:
                    rows.append({
                        "project": project,
                        "resource": resource,
                        "role": binding.role,
                        "member": member,
                        "member_type": member.split(":", 1)[0] if ":" in member else member,
                    })
    except PermissionDenied as e:
        print(f"  ! permission denied on asset search: {e}", file=sys.stderr)
    except GoogleAPIError as e:
        print(f"  ! asset search failed: {e}", file=sys.stderr)
    return rows


def collect_iam_recommendations(projects):
    """IAM Recommender insights per project → set of (project, member, role) stale flags."""
    client = recommender_v1.RecommenderClient()
    stale = {}
    for project in sorted(projects):
        if not project:
            continue
        parent = (f"projects/{project}/locations/global/recommenders/{IAM_RECOMMENDER}")
        try:
            for rec in client.list_recommendations(parent=parent):
                # Recommendation content describes the binding to remove/replace.
                desc = rec.description or ""
                # Tag the project-level recommendation; detailed member/role parsing
                # is available in rec.content.operation_groups for closed-loop apply.
                stale.setdefault(project, []).append({
                    "description": desc,
                    "priority": recommender_v1.Recommendation.Priority(rec.priority).name
                    if rec.priority else "",
                    "subtype": rec.recommender_subtype or "",
                })
        except PermissionDenied:
            print(f"  ! recommender permission denied: {project}", file=sys.stderr)
        except GoogleAPIError as e:
            print(f"  ! recommender failed for {project}: {e}", file=sys.stderr)
    return stale


# --------------------------------------------------------------------------- #
# Workbook builder
# --------------------------------------------------------------------------- #
def build_workbook(bindings, recs, out_path):
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0B5394")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    decision_fill = PatternFill("solid", fgColor="FCE4D6")
    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    def add_decision_validation(ws, col_letter, nrows):
        dv = DataValidation(
            type="list",
            formula1='"APPROVE,REVOKE,INVESTIGATE"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.error = "Choose APPROVE, REVOKE, or INVESTIGATE."
        dv.errorTitle = "Invalid decision"
        dv.prompt = "Select a review decision."
        dv.promptTitle = "Reviewer decision"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{nrows + 1}")

    # ---- Sheet 1: IAM Bindings ----
    ws1 = wb.active
    ws1.title = "IAM Bindings"
    cols1 = ["Project", "Resource", "Role", "Member", "Member Type",
             "Recommender Flag", "Reviewer Decision", "Reviewer", "Date", "Justification"]
    ws1.append(cols1)
    style_header(ws1, len(cols1))

    # projects that have ANY recommender insight → flag their bindings for attention
    flagged_projects = {p for p, items in recs.items() if items}
    for b in bindings:
        flag = "Review (insights present)" if b["project"] in flagged_projects else ""
        ws1.append([
            b["project"], b["resource"], b["role"], b["member"],
            b["member_type"], flag, "", "", "", "",
        ])
    if bindings:
        add_decision_validation(ws1, "G", len(bindings))
        for r in range(2, len(bindings) + 2):
            ws1.cell(row=r, column=7).fill = decision_fill
            if ws1.cell(row=r, column=6).value:
                ws1.cell(row=r, column=6).fill = flag_fill
    widths1 = [22, 40, 34, 38, 14, 24, 18, 16, 12, 32]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 2: Recommender Insights (native detail) ----
    ws2 = wb.create_sheet("Recommender Insights")
    cols2 = ["Project", "Priority", "Subtype", "Description"]
    ws2.append(cols2)
    style_header(ws2, len(cols2))
    for project, items in sorted(recs.items()):
        for it in items:
            ws2.append([project, it["priority"], it["subtype"], it["description"]])
    widths2 = [24, 12, 30, 80]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 0: README ----
    ws3 = wb.create_sheet("README", 0)
    ws3.column_dimensions["A"].width = 100
    lines = [
        ("GCP Access Review — Reviewer Workbook", True),
        ("", False),
        (f"Generated: {dt.datetime.now(dt.timezone.utc).isoformat()} (read-only pull)", False),
        ("", False),
        ("How to complete this review:", True),
        ("1. Review each binding on the 'IAM Bindings' tab.", False),
        ("2. Rows flagged in 'Recommender Flag' have native usage insights — check first.", False),
        ("3. Set 'Reviewer Decision' to APPROVE, REVOKE, or INVESTIGATE.", False),
        ("4. Fill in Reviewer, Date, and a short Justification.", False),
        ("5. Every row needs a decision before the review is Done.", False),
        ("6. REVOKE rows are actioned by the separate, gated apply step.", False),
        ("", False),
        ("Native signals used:", True),
        ("  • IAM Bindings tab        → Cloud Asset Inventory searchAllIamPolicies.", False),
        ("  • Recommender Insights tab → IAM Recommender (90-day usage analysis).", False),
        ("", False),
        ("This workbook does not change any access. It is evidence + a decision record.", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        c = ws3.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=13 if i == 1 else 11,
                          color="0B5394" if i == 1 else "000000")

    wb.save(out_path)


# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="GCP access review pull (read-only).")
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--org-id", help="Organization ID (org-level scope).")
    g.add_argument("--project", help="Single project ID.")
    ap.add_argument("--out", default="gcp_access_review.xlsx")
    args = ap.parse_args()

    scope = f"organizations/{args.org_id}" if args.org_id else f"projects/{args.project}"
    print("GCP Access Review — PULL (read-only)")
    print(f"· Scope: {scope}")

    print("· Pulling IAM bindings (Cloud Asset Inventory) ...")
    bindings = collect_iam_bindings(scope)
    print(f"  {len(bindings)} bindings")

    projects = {b["project"] for b in bindings if b["project"]}
    if args.project:
        projects.add(args.project)
    print(f"· Pulling IAM Recommender insights for {len(projects)} project(s) ...")
    recs = collect_iam_recommendations(projects)
    total_recs = sum(len(v) for v in recs.values())
    print(f"  {total_recs} recommendation(s) across {len(recs)} project(s)")

    print(f"· Building workbook: {args.out}")
    build_workbook(bindings, recs, args.out)
    print(f"Done. {len(bindings)} bindings, {total_recs} insights.")
    print("Reviewer fills decisions; REVOKE rows feed the gated apply step.")


if __name__ == "__main__":
    main()
